from urllib import request
from openpyxl import load_workbook
import sqlite3
import os
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
import io
import pypdn
import numpy as np
from PIL import Image


def xlsx_downloader():
    # удалим старый xlsx файл
    xlsx = "Банк диалектологических карт.xlsx"
    if xlsx in os.listdir():
        os.remove(xlsx)
        print('Старый xlsx файл успешно удалён')
    else:
        print('Старый xlsx файл не обнаружен')

    # скачаем новый xlsx файл
    sheet_id = "1AhZbjYB3AfbzB2JoclQKvdBEYQbCXbuVZ8pf2hbsEBk"
    url = "https://docs.google.com/spreadsheets/export?id=" + sheet_id
    request.urlretrieve(url, xlsx)
    print('Новый xlsx файл успешно скачан')


def xlsx_to_db():
    # удалим старый db файл
    db = "Банк диалектологических карт"
    if f"{db}.db" in os.listdir():
        os.remove(f"{db}.db")
        print('Старый db файл успешно удалён')
    else:
        print('Старый db файл не обнаружен')

    # прочитаем xlsx файл
    wb = load_workbook(f"{db}.xlsx")

    # создадим db файл
    open(f"{db}.db", 'a').close()
    print('Новый db файл успешно создан')

    # получим список листов
    sheets = [_ for _ in wb.sheetnames if " (копия)" not in _]

    # откроем первый лист
    sheet = wb[sheets[0]]

    # соберём строки
    raw_rows = []
    for row in sheet:
        if not all([cell.value is None for cell in row]):
            raw_rows.append([cell.value for cell in row if cell.value is not None])

    # получим имена колонок и данные карт
    column_names = raw_rows[0]
    rows = [_ for _ in raw_rows[1:] if (set(_) & {'сверена', 'готова'})]

    # запишем главную таблицу в sql
    conn = sqlite3.connect(f"{db}.db")
    cur = conn.cursor()
    create = ''
    column_types = ["INTEGER" if any(_ in column for _ in ["№", "число"]) else "TEXT" for column in column_names]
    for column, column_type in zip(column_names, column_types):
        create += f'"{column}" {column_type},\n'
    cmd = f'CREATE TABLE IF NOT EXISTS "{sheets[0]}" (\n{create[:-2]})'
    cur.execute(cmd)

    for row in rows:
        insert = f"""INSERT INTO "{sheets[0]}" ({str(column_names)[1:-1]}) VALUES ({str(row)[1:-1]})"""
        cur.execute(insert)

    print('Общая таблица успешно перенесена в db файл')

    # перенесём таблицы признаков
    for sheet_name in sheets[1:]:
        sheet = wb[sheet_name]

        try:
            raw_rows = [[int(cell.value) if isinstance(cell.value, float) else cell.value for cell in row]
                        for row in sheet.iter_rows() if any(cell.value for cell in row)]

            column_names = []
            for column_name in raw_rows[0]:
                if column_name:
                    if column_name not in column_names:
                        column_names.append(column_name)
                    else:
                        if not column_name[-1].isdigit():
                            column_names.append(column_name + "_0")
                        else:
                            column_names.append(column_name[:-1] + str(int(column_name[-1]) + 1))

            length = len(column_names)
            rows = [[value if value else "" for value in row[:length]] for row in raw_rows[1:]]

            create = ''
            column_types = ["INTEGER" if any(_ in column for _ in ["№", "число"]) else "TEXT" for column in column_names]
            for column, column_type in zip(column_names, column_types):
                create += f'"{column}" {column_type},\n'
            cmd = f'CREATE TABLE IF NOT EXISTS "{sheet_name}" (\n{create[:-2]})'
            cur.execute(cmd)

            for row in rows:
                values = str(row)[1:-1].replace("\\n", "\n")
                insert = f"""INSERT INTO "{sheet_name}" ({str(column_names)[1:-1]}) VALUES ({values})"""
                insert = insert.replace("\\'", "''")
                cur.execute(insert)

            print(f'Таблица {sheet_name} успешно перенесена в db файл')

        except Exception as e:
            print(f'Таблицы {sheet_name} не была перенесена в db файл. Возникла ошибка {e}')

    conn.commit()
    conn.close()
    os.remove(f"{db}.xlsx")
    print('Все таблицы успешно перенесены в db файл')
    print('Новый db файл успешно создан')


def remove_imgs():
    directory = "maps"
    if directory in os.listdir():
        for entity in os.listdir(directory):
            if all(format not in entity for format in [".pdn", ".png", ".json"]):
                for layer in os.listdir(f"{directory}/{entity}"):
                    os.remove(f"{directory}/{entity}/{layer}")
                os.rmdir(f"{directory}/{entity}")
            else:
                os.remove(f"{directory}/{entity}")
        os.rmdir(directory)
        print('Все старые файлы изображений успешно удалены')


def download_pdns():
    scopes = ['https://www.googleapis.com/auth/drive']
    service_account_file = 'dialectography-credentials.json'
    credentials = service_account.Credentials.from_service_account_file(service_account_file, scopes=scopes)
    service = build('drive', 'v3', credentials=credentials)

    def folder_download(folder_id):
        results = service.files().list(pageSize=1000,
                                       fields="nextPageToken, files(id, name, mimeType)",
                                       q=f"'{folder_id}' in parents"
                                       ).execute()

        for file in results['files']:
            if file['mimeType'] == 'application/vnd.google-apps.folder':
                path = file['name']
                if path not in os.listdir():
                    os.mkdir(path)

                os.chdir(path)
                folder_download(file["id"])
                os.chdir(os.pardir)
            else:
                request = service.files().get_media(fileId=file["id"])
                fh = io.FileIO(file['name'], 'wb')
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                    if file["name"].endswith(".pdn"):
                        print(f'{file["name"]} успешно скачана')

    path = "maps"
    if path not in os.listdir():
        os.mkdir(path)

    os.chdir(path)
    folder_id = '1Dk7U1Xk8R2_8QliAnsM0HfjTpHR3lNE4'
    folder_download(folder_id)
    os.chdir(os.pardir)

    print('Все pdn файлы успешно скачаны')


def pdn_to_bins():
    for file in os.listdir("maps"):
        if file.endswith(".pdn"):
            layered_image = pypdn.read(f"maps/{file}")

            directory = file.rstrip(".pdn")
            if directory not in os.listdir("maps"):
                os.mkdir(f"maps/{directory}")

            if directory != "База":
                for layer in layered_image.layers:
                    img = layer.image[:, :, 3] != 0
                    with open(f"maps/{directory}/{layer.name}.bin", 'wb') as f:
                        f.write(np.packbits(img))
            else:
                for layer in layered_image.layers:
                    img = Image.fromarray(layer.image)
                    img.save(f"maps/{directory}/{layer.name}.png")

            os.remove(f"maps/{file}")
            print(f'{file} успешно разбита на bin файлы')

    print('Все pdn файлы успешно разбиты на bin файлы')